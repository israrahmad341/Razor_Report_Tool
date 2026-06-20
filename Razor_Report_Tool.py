import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- SQL Server connection details (left as is for local environment) ---
server_ip = r'DESKTOP-U3MJ8IT\\SQLEXPRESS'
database = 'SPAYINDIADB'
driver = 'ODBC Driver 17 for SQL Server'


def generate_report(razor_orders_file: str, payments_report_file: str, asm_file: str, output_dir: str) -> bool:
    """
    Generates the Razor Report by processing RazorOrders and Payments data.

    Args:
        razor_orders_file (str): Path to the RazorOrders Excel file.
        payments_report_file (str): Path to the Payments_Report Excel file.
        asm_file (str): Path to the ASM mapping file.
        output_dir (str): Directory where the final report will be saved.

    Returns:
        bool: True if successful, False otherwise.
    """
    try:
        print("--- Starting Razor Report Generation ---")

        # 1) Load the ASM data
        print(f"Reading ASM file: {asm_file}")
        asm_df = pd.read_excel(asm_file)
        required_asm_columns = ['code', 'head', 'asmname', 'parentcode']
        missing_asm_columns = [col for col in required_asm_columns if col not in asm_df.columns]
        if missing_asm_columns:
            print(f"❌ ERROR in Razor Report: Missing required columns in {os.path.basename(asm_file)}: {missing_asm_columns}")
            return False

        # 2) Load RazorOrders file
        print(f"Reading RazorOrders file: {razor_orders_file}")
        razor_df = pd.read_excel(razor_orders_file)
        required_razor_columns = ['RequestNo', 'code', 'Retailer']
        missing_razor_columns = [col for col in required_razor_columns if col not in razor_df.columns]
        if missing_razor_columns:
            print(f"❌ ERROR in Razor Report: Missing required columns in {os.path.basename(razor_orders_file)}: {missing_razor_columns}")
            return False

        # Extract report date (prefer date from RazorOrders if present)
        report_date = None
        addeddate_col = next((col for col in razor_df.columns if col.lower() == 'addeddate'), None)
        if addeddate_col:
            razor_df[addeddate_col] = pd.to_datetime(razor_df[addeddate_col], errors='coerce')
            non_null_dates = razor_df[addeddate_col].dropna()
            if not non_null_dates.empty:
                report_date = non_null_dates.iloc[0].strftime('%d %b %Y').upper()
        if not report_date:
            from datetime import datetime
            report_date = datetime.now().strftime('%d %b %Y').upper()
            print(f"⚠️ Could not determine date from RazorOrders, using current date: {report_date}")

        # 3) Load Payments_Report file
        print(f"Reading Payments Report file: {payments_report_file}")
        payments_df = pd.read_excel(payments_report_file)
        required_payments_columns = ['Order Id', 'Status', 'Total Amount']
        missing_payments_columns = [col for col in required_payments_columns if col not in payments_df.columns]
        if missing_payments_columns:
            print(f"❌ ERROR in Razor Report: Missing required columns in {os.path.basename(payments_report_file)}: {missing_payments_columns}")
            return False

        # ---- Normalize/Process data ----

        # Find case-insensitive columns
        status_col = next((col for col in razor_df.columns if col.lower() == 'status'), None)
        pgtype_col = next((col for col in razor_df.columns if col.lower() == 'pgtype'), None)
        amount_col = next((col for col in razor_df.columns if col.lower() == 'amount'), None)

        if not all([status_col, pgtype_col, amount_col]):
            print("❌ ERROR in Razor Report: 'status', 'PGType', or 'amount' column not found in RazorOrders file.")
            return False

        pg2_mask = razor_df[pgtype_col] == 'PG-2'
        captured_mask = razor_df[status_col].fillna('').str.lower() == 'captured'
        razor_df_all = razor_df[pg2_mask & captured_mask].copy()
        razor_df_all = razor_df_all.rename(columns={status_col: 'status', pgtype_col: 'PGType', amount_col: 'amount'})
        razor_df_all['amount'] = pd.to_numeric(razor_df_all['amount'], errors='coerce').fillna(0)

        # Payments processing
        payment_status_col = next((col for col in payments_df.columns if col.lower() == 'status'), None)
        total_amount_col = next((col for col in payments_df.columns if 'total' in col.lower() and 'amount' in col.lower()), None)
        if not total_amount_col:  # fallback
            total_amount_col = next((col for col in payments_df.columns if 'amount' in col.lower()), None)

        if not all([payment_status_col, total_amount_col]):
            print("❌ ERROR in Razor Report: 'Status' or 'Total Amount' column not found in Payments_Report file.")
            return False

        payments_df_temp = payments_df
        if 'PGType' in payments_df.columns:
            payments_df_temp = payments_df[payments_df['PGType'] == 'PG-5'].copy()

        payments_df_accepted = payments_df_temp[payments_df_temp[payment_status_col].str.lower() == 'accepted'].copy()
        payments_df_accepted = payments_df_accepted.rename(columns={payment_status_col: 'Status', total_amount_col: 'Total Amount'})
        payments_df_accepted['Total Amount'] = pd.to_numeric(payments_df_accepted['Total Amount'], errors='coerce').fillna(0)

        # Process PG-2 and PG-5 data
        razor_with_asm = pd.merge(razor_df_all, asm_df[['code', 'head', 'asmname']], on='code', how='left')
        razor_with_asm_mapped = razor_with_asm[razor_with_asm['head'].notna()].copy()

        razor_df_full = razor_df.rename(columns={status_col: 'status', pgtype_col: 'PGType', amount_col: 'amount'})
        payments_with_code = pd.merge(payments_df_accepted, razor_df_full[['RequestNo', 'code', 'Retailer']], left_on='Order Id', right_on='RequestNo', how='left')
        payments_with_asm = pd.merge(payments_with_code, asm_df[['code', 'head', 'asmname']], on='code', how='left')
        payments_with_asm_mapped = payments_with_asm[payments_with_asm['head'].notna()].copy()
        payments_without_asm = payments_with_asm[payments_with_asm['head'].isna()].copy()
        unmatched_payments_amount = payments_without_asm['Total Amount'].sum() if not payments_without_asm.empty else 0

        if razor_with_asm_mapped.empty and payments_with_asm_mapped.empty:
            print("⚠️ No transaction records with ASM mapping found. Report will be empty.")
            final_report_df = pd.DataFrame(columns=['Head', 'ASM', 'FTD PG-2', 'Outlet Count PG-2', 'FTD PG-5', 'Outlet Count PG-5', 'Total PG', 'Total Outlet Count'])
            output_rows = []
        else:
            pg2_grouped = pd.DataFrame(columns=['head', 'asmname', 'FTD_PG2', 'Outlet_Count_PG2'])
            if not razor_with_asm_mapped.empty:
                pg2_grouped = razor_with_asm_mapped.groupby(['head', 'asmname']).agg(FTD_PG2=('amount', 'sum'), Outlet_Count_PG2=('Retailer', 'nunique')).reset_index()

            pg5_grouped = pd.DataFrame(columns=['head', 'asmname', 'FTD_PG5', 'Outlet_Count_PG5'])
            if not payments_with_asm_mapped.empty:
                pg5_grouped = payments_with_asm_mapped.groupby(['head', 'asmname']).agg(FTD_PG5=('Total Amount', 'sum'), Outlet_Count_PG5=('Retailer', lambda x: x.dropna().nunique())).reset_index()

            final_report_df = pd.merge(pg2_grouped, pg5_grouped, on=['head', 'asmname'], how='outer').fillna(0)

            all_retailers = pd.concat([
                razor_with_asm_mapped[['head', 'asmname', 'Retailer']].dropna() if not razor_with_asm_mapped.empty else pd.DataFrame(columns=['head', 'asmname', 'Retailer']),
                payments_with_asm_mapped[['head', 'asmname', 'Retailer']].dropna() if not payments_with_asm_mapped.empty else pd.DataFrame(columns=['head', 'asmname', 'Retailer'])
            ], ignore_index=True)

            if not all_retailers.empty:
                unique_counts = all_retailers.groupby(['head', 'asmname'])['Retailer'].nunique().reset_index()
                unique_counts.columns = ['head', 'asmname', 'Unique_Outlet_Count']
                final_report_df = pd.merge(final_report_df, unique_counts, on=['head', 'asmname'], how='left')
            else:
                final_report_df['Unique_Outlet_Count'] = 0

            final_report_df['Unique_Outlet_Count'] = final_report_df['Unique_Outlet_Count'].fillna(0)

            all_asms = asm_df[['head', 'asmname']].drop_duplicates().dropna()
            final_report_df = pd.merge(all_asms, final_report_df, on=['head', 'asmname'], how='left').fillna(0)

            final_report_df['Grand_Total'] = final_report_df.get('FTD_PG2', 0) + final_report_df.get('FTD_PG5', 0)
            final_report_df['Total_Outlet_Count'] = final_report_df['Unique_Outlet_Count']

            final_report_df = final_report_df.rename(columns={'head': 'Head', 'asmname': 'ASM', 'FTD_PG2': 'FTD PG-2', 'Outlet_Count_PG2': 'Outlet Count PG-2', 'FTD_PG5': 'FTD PG-5', 'Outlet_Count_PG5': 'Outlet Count PG-5', 'Grand_Total': 'Total PG', 'Total_Outlet_Count': 'Total Outlet Count'})

            if 'Unique_Outlet_Count' in final_report_df.columns:
                final_report_df = final_report_df.drop(columns=['Unique_Outlet_Count'])

            final_report_df = final_report_df[(final_report_df['FTD PG-2'] > 0) | (final_report_df['FTD PG-5'] > 0)].copy()
            final_report_df = final_report_df.sort_values(by=['Head', 'ASM']).reset_index(drop=True)

            output_rows = []
            for head_name, head_group_df in final_report_df.groupby('Head'):
                output_rows.append(head_group_df)
                subtotal_row = pd.DataFrame({'Head': [f'{head_name} Total'], 'ASM': [''], 'FTD PG-2': [head_group_df['FTD PG-2'].sum()], 'Outlet Count PG-2': [head_group_df['Outlet Count PG-2'].sum()], 'FTD PG-5': [head_group_df['FTD PG-5'].sum()], 'Outlet Count PG-5': [head_group_df['Outlet Count PG-5'].sum()], 'Total PG': [head_group_df['Total PG'].sum()], 'Total Outlet Count': [head_group_df['Total Outlet Count'].sum()]})
                output_rows.append(subtotal_row)

            if not final_report_df.empty:
                grand_total_pg2 = final_report_df['FTD PG-2'].sum()
                grand_total_outlet_pg2 = final_report_df['Outlet Count PG-2'].sum()
                grand_total_pg5 = final_report_df['FTD PG-5'].sum() + unmatched_payments_amount
                grand_total_outlet_pg5 = final_report_df['Outlet Count PG-5'].sum()
                grand_total_combined = grand_total_pg2 + grand_total_pg5
                grand_total_outlet_combined = final_report_df['Total Outlet Count'].sum()
                grand_total_row = pd.DataFrame({'Head': ['Grand Total'], 'ASM': [''], 'FTD PG-2': [grand_total_pg2], 'Outlet Count PG-2': [grand_total_outlet_pg2], 'FTD PG-5': [grand_total_pg5], 'Outlet Count PG-5': [grand_total_outlet_pg5], 'Total PG': [grand_total_combined], 'Total Outlet Count': [grand_total_outlet_combined]})
                output_rows.append(grand_total_row)

        # Save to Excel
        output_filename = 'ASM_WISE_PG_BUSINESS_WITH_MERCHANT_COUNT.xlsx'
        full_output_path = os.path.join(output_dir, output_filename)

        excel_df = pd.concat(output_rows, ignore_index=True) if output_rows else final_report_df

        with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
            excel_df.to_excel(writer, index=False, sheet_name='ASM WISE PG BUSINESS', startrow=1)
            workbook = writer.book
            worksheet = writer.sheets['ASM WISE PG BUSINESS']

            # Styling
            worksheet['A1'] = f'ASM WISE PG BUSINESS WITH MERCHANT COUNT {report_date}'
            worksheet.merge_cells('A1:H1')
            title_font = Font(color='FFFFFF', bold=True, size=12)
            title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            worksheet['A1'].font = title_font
            worksheet['A1'].fill = title_fill
            worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

            header_font = Font(color='000000', bold=True)
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for col in range(1, 9):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx in range(3, worksheet.max_row + 1):
                head_value = worksheet.cell(row=row_idx, column=1).value
                is_grand_total = head_value == 'Grand Total'
                is_subtotal = head_value and 'Total' in str(head_value) and not is_grand_total

                fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                font = Font()
                if is_grand_total:
                    fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    font = Font(color='FFFFFF', bold=True)
                elif is_subtotal:
                    fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                    font = Font(bold=True)

                for col in range(1, 9):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.fill = fill
                    cell.font = font

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=8):
                for cell in row:
                    cell.border = thin_border

            for col_idx in range(1, 9):
                max_length = max(len(str(cell.value)) for cell in worksheet[get_column_letter(col_idx)] if cell.value)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        print(f"✅ Razor Report saved successfully at: {full_output_path}")
        print("--- Finished Razor Report Generation ---")
        return True

    except FileNotFoundError as e:
        print(f"❌ ERROR in Razor Report: File not found - {e.filename}. Please check file names in the 'source' directory.")
        return False
    except KeyError as e:
        print(f"❌ ERROR in Razor Report: A required column is missing from an Excel file: {e}. Please check your input files.")
        return False
    except Exception as e:
        print(f"❌ An unexpected error occurred in Razor Report: {e}")
        return False


if __name__ == "__main__":
    print("This script is designed to be imported as a module. Run master_tool.py to generate reports.")